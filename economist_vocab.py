#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import random
import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Iterable

import openpyxl


DEFAULT_DB_PATH = Path("economist_vocab.db")
DEFAULT_WORKBOOKS = [
    "/Users/lawrencecheng/Desktop/The Economist/50~99 (3924).xlsx",
    "/Users/lawrencecheng/Desktop/The Economist/100~199 (3180).xlsx",
    "/Users/lawrencecheng/Desktop/The Economist/200~499 (3176).xlsx",
    "/Users/lawrencecheng/Desktop/The Economist/500~1999 (3000).xlsx",
    "/Users/lawrencecheng/Desktop/The Economist/2000~ (2330).xlsx",
]

UTC = timezone.utc


@dataclass(frozen=True)
class BandInfo:
    label: str
    rank: int


SCHEMA = """
CREATE TABLE IF NOT EXISTS words (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    lemma TEXT NOT NULL,
    normalized_lemma TEXT NOT NULL UNIQUE,
    best_band_label TEXT NOT NULL,
    best_band_rank INTEGER NOT NULL,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS source_entries (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    word_id INTEGER NOT NULL REFERENCES words(id) ON DELETE CASCADE,
    workbook_name TEXT NOT NULL,
    sheet_name TEXT NOT NULL,
    row_number INTEGER NOT NULL,
    band_label TEXT NOT NULL,
    band_rank INTEGER NOT NULL,
    pos TEXT,
    meanings_json TEXT NOT NULL,
    extra_json TEXT NOT NULL,
    source_signature TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS study_cards (
    word_id INTEGER PRIMARY KEY REFERENCES words(id) ON DELETE CASCADE,
    status TEXT NOT NULL DEFAULT 'new',
    notes TEXT NOT NULL DEFAULT '',
    correct_count INTEGER NOT NULL DEFAULT 0,
    wrong_count INTEGER NOT NULL DEFAULT 0,
    streak INTEGER NOT NULL DEFAULT 0,
    ease REAL NOT NULL DEFAULT 2.5,
    interval_days REAL NOT NULL DEFAULT 0,
    last_reviewed_at TEXT,
    next_review_at TEXT,
    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS review_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    word_id INTEGER NOT NULL REFERENCES words(id) ON DELETE CASCADE,
    reviewed_at TEXT NOT NULL,
    prompt_mode TEXT NOT NULL,
    grade TEXT NOT NULL
);
"""


def utc_now() -> datetime:
    return datetime.now(UTC)


def connect(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.executescript(SCHEMA)
    return conn


def normalize_word(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def parse_band(path: Path) -> BandInfo:
    stem = path.stem
    match = re.match(r"(\d+)~", stem)
    if match:
        return BandInfo(label=stem, rank=int(match.group(1)))
    raise ValueError(f"Unable to infer frequency band from {path.name}")


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def iter_workbook_entries(path: Path) -> Iterable[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    band = parse_band(path)
    for sheet_name in wb.sheetnames:
        if sheet_name.lower().startswith("total"):
            continue
        ws = wb[sheet_name]
        header_mode = False
        header_indices: dict[str, int] = {}
        for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not row:
                continue
            if row_number == 1:
                first = normalize_header(row[0]) if row else ""
                second = normalize_header(row[1]) if len(row) > 1 else ""
                first_ok = first in {"vocabulary", "word"}
                second_ok = second in {"type of word", "type", "part of speech"}
                if first_ok and second_ok:
                    header_mode = True
                    english_headers = {"english definition", "english definitions", "definition in english"}
                    chinese_headers = {"chinese definition", "chinese definitions", "中文定義", "中文定义", "中文释义", "中文釋義"}
                    simplified_chinese_headers = {
                        "simplified chinese definition",
                        "simplified chinese definitions",
                        "simplified chinese",
                        "简体中文定义",
                        "簡體中文定義",
                        "简体中文释义",
                        "簡體中文釋義",
                    }
                    example_headers = {"example sentence", "example", "example sentences"}
                    pronunciation_headers = {"pronunciation", "ipa", "ipa pronunciation", "international phonetic alphabet"}
                    for index, cell in enumerate(row):
                        header = normalize_header(cell)
                        if header in english_headers:
                            header_indices["english_definition"] = index
                        elif header in chinese_headers:
                            header_indices["chinese_definition"] = index
                        elif header in simplified_chinese_headers:
                            header_indices["simplified_chinese_definition"] = index
                        elif header in example_headers:
                            header_indices["example_sentence"] = index
                        elif header in pronunciation_headers:
                            header_indices["pronunciation"] = index
                    continue
            raw_word = row[0]
            if raw_word is None or not isinstance(raw_word, str):
                continue
            lemma = raw_word.strip()
            if not lemma:
                continue
            pos = row[1] if len(row) > 1 and isinstance(row[1], str) else None
            english_definition = ""
            example_sentence = ""
            pronunciation = ""
            if header_mode:
                english_index = header_indices.get("english_definition")
                chinese_index = header_indices.get("chinese_definition")
                simplified_chinese_index = header_indices.get("simplified_chinese_definition")
                example_index = header_indices.get("example_sentence")
                pronunciation_index = header_indices.get("pronunciation")
                english_definition = str(row[english_index]).strip() if english_index is not None and english_index < len(row) and row[english_index] is not None else ""
                chinese_definition = str(row[chinese_index]).strip() if chinese_index is not None and chinese_index < len(row) and row[chinese_index] is not None else ""
                simplified_chinese_definition = str(row[simplified_chinese_index]).strip() if simplified_chinese_index is not None and simplified_chinese_index < len(row) and row[simplified_chinese_index] is not None else ""
                example_sentence = str(row[example_index]).strip() if example_index is not None and example_index < len(row) and row[example_index] is not None else ""
                pronunciation = str(row[pronunciation_index]).strip() if pronunciation_index is not None and pronunciation_index < len(row) and row[pronunciation_index] is not None else ""
                meanings = [item.strip() for item in chinese_definition.splitlines() if item and item.strip()]
                if not meanings and chinese_definition:
                    meanings = [chinese_definition]
            else:
                extras = []
                for cell in row[2:]:
                    if cell is None:
                        continue
                    text = str(cell).strip()
                    if text:
                        extras.append(text)
                meanings = extras[:]
            signature = f"{path.name}|{sheet_name}|{row_number}|{normalize_word(lemma)}"
            yield {
                "lemma": lemma,
                "normalized_lemma": normalize_word(lemma),
                "band_label": band.label,
                "band_rank": band.rank,
                "workbook_name": path.name,
                "sheet_name": sheet_name,
                "row_number": row_number,
                "pos": pos,
                "meanings_json": json.dumps(meanings, ensure_ascii=False),
                "extra_json": json.dumps(
                    {
                        "raw_cells": [None if c is None else str(c) for c in row[2:]],
                        "english_definition": english_definition,
                        "simplified_chinese_definition": simplified_chinese_definition,
                        "example_sentence": example_sentence,
                        "pronunciation": pronunciation,
                        "header_mode": header_mode,
                    },
                    ensure_ascii=False,
                ),
                "source_signature": signature,
            }


def ensure_card(conn: sqlite3.Connection, word_id: int) -> None:
    conn.execute(
        """
        INSERT INTO study_cards (word_id)
        VALUES (?)
        ON CONFLICT(word_id) DO NOTHING
        """,
        (word_id,),
    )


def import_workbooks(conn: sqlite3.Connection, workbook_paths: list[Path], reset: bool) -> dict:
    stats = {"inserted_words": 0, "inserted_entries": 0, "updated_best_band": 0}
    if reset:
        conn.executescript(
            """
            DELETE FROM review_log;
            DELETE FROM source_entries;
            DELETE FROM study_cards;
            DELETE FROM words;
            """
        )
    for workbook_path in workbook_paths:
        for entry in iter_workbook_entries(workbook_path):
            row = conn.execute(
                "SELECT id, best_band_rank FROM words WHERE normalized_lemma = ?",
                (entry["normalized_lemma"],),
            ).fetchone()
            if row is None:
                cursor = conn.execute(
                    """
                    INSERT INTO words (lemma, normalized_lemma, best_band_label, best_band_rank)
                    VALUES (?, ?, ?, ?)
                    """,
                    (
                        entry["lemma"],
                        entry["normalized_lemma"],
                        entry["band_label"],
                        entry["band_rank"],
                    ),
                )
                word_id = cursor.lastrowid
                stats["inserted_words"] += 1
            else:
                word_id = row["id"]
                if entry["band_rank"] < row["best_band_rank"]:
                    conn.execute(
                        """
                        UPDATE words
                        SET lemma = ?, best_band_label = ?, best_band_rank = ?
                        WHERE id = ?
                        """,
                        (entry["lemma"], entry["band_label"], entry["band_rank"], word_id),
                    )
                    stats["updated_best_band"] += 1
            ensure_card(conn, word_id)
            cursor = conn.execute(
                """
                INSERT INTO source_entries (
                    word_id, workbook_name, sheet_name, row_number, band_label, band_rank,
                    pos, meanings_json, extra_json, source_signature
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(source_signature) DO UPDATE SET
                    word_id = excluded.word_id,
                    workbook_name = excluded.workbook_name,
                    sheet_name = excluded.sheet_name,
                    row_number = excluded.row_number,
                    band_label = excluded.band_label,
                    band_rank = excluded.band_rank,
                    pos = excluded.pos,
                    meanings_json = excluded.meanings_json,
                    extra_json = excluded.extra_json
                """,
                (
                    word_id,
                    entry["workbook_name"],
                    entry["sheet_name"],
                    entry["row_number"],
                    entry["band_label"],
                    entry["band_rank"],
                    entry["pos"],
                    entry["meanings_json"],
                    entry["extra_json"],
                    entry["source_signature"],
                ),
            )
            stats["inserted_entries"] += cursor.rowcount
    conn.commit()
    return stats


def refresh_workbooks(conn: sqlite3.Connection, workbook_paths: list[Path]) -> dict:
    stats = {
        "deleted_entries": 0,
        "inserted_words": 0,
        "inserted_entries": 0,
        "updated_best_band": 0,
        "deleted_orphan_words": 0,
    }
    workbook_names = [path.name for path in workbook_paths]
    placeholder = ", ".join("?" for _ in workbook_names)

    deleted_entries = conn.execute(
        f"DELETE FROM source_entries WHERE workbook_name IN ({placeholder})",
        workbook_names,
    ).rowcount
    stats["deleted_entries"] = deleted_entries

    import_stats = import_workbooks(conn, workbook_paths, reset=False)
    stats["inserted_words"] = import_stats["inserted_words"]
    stats["inserted_entries"] = import_stats["inserted_entries"]
    stats["updated_best_band"] = import_stats["updated_best_band"]

    conn.execute(
        """
        UPDATE words
        SET best_band_label = (
                SELECT band_label
                FROM source_entries
                WHERE source_entries.word_id = words.id
                ORDER BY band_rank, workbook_name, row_number
                LIMIT 1
            ),
            best_band_rank = (
                SELECT band_rank
                FROM source_entries
                WHERE source_entries.word_id = words.id
                ORDER BY band_rank, workbook_name, row_number
                LIMIT 1
            )
        WHERE EXISTS (
            SELECT 1
            FROM source_entries
            WHERE source_entries.word_id = words.id
        )
        """
    )

    orphan_rows = conn.execute(
        """
        SELECT id
        FROM words
        WHERE NOT EXISTS (
            SELECT 1
            FROM source_entries
            WHERE source_entries.word_id = words.id
        )
        """
    ).fetchall()
    if orphan_rows:
        orphan_ids = [row["id"] for row in orphan_rows]
        orphan_placeholder = ", ".join("?" for _ in orphan_ids)
        conn.execute(f"DELETE FROM words WHERE id IN ({orphan_placeholder})", orphan_ids)
        stats["deleted_orphan_words"] = len(orphan_ids)

    conn.commit()
    return stats


def print_stats(conn: sqlite3.Connection) -> None:
    totals = conn.execute(
        """
        SELECT
            COUNT(*) AS words,
            SUM(CASE WHEN status = 'new' THEN 1 ELSE 0 END) AS new_cards,
            SUM(CASE WHEN next_review_at IS NOT NULL AND next_review_at <= ? THEN 1 ELSE 0 END) AS due_cards
        FROM words
        JOIN study_cards ON study_cards.word_id = words.id
        """,
        (utc_now().isoformat(),),
    ).fetchone()
    print(f"Words: {totals['words']}")
    print(f"New cards: {totals['new_cards']}")
    print(f"Due now: {totals['due_cards']}")
    print("\nBy frequency band:")
    for row in conn.execute(
        """
        SELECT best_band_label, COUNT(*) AS total
        FROM words
        GROUP BY best_band_label, best_band_rank
        ORDER BY best_band_rank
        """
    ):
        print(f"  {row['best_band_label']}: {row['total']}")


def lookup_word(conn: sqlite3.Connection, term: str) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT id, lemma, normalized_lemma, best_band_label, best_band_rank
        FROM words
        WHERE normalized_lemma = ?
        """,
        (normalize_word(term),),
    ).fetchone()


def show_word(conn: sqlite3.Connection, term: str) -> int:
    word = lookup_word(conn, term)
    if word is None:
        print(f"No word found for: {term}")
        return 1
    card = conn.execute("SELECT * FROM study_cards WHERE word_id = ?", (word["id"],)).fetchone()
    print(f"{word['lemma']}  [{word['best_band_label']}]")
    if card["notes"]:
        print(f"Notes: {card['notes']}")
    if card["last_reviewed_at"]:
        print(
            f"Review: status={card['status']} streak={card['streak']} next={card['next_review_at'] or 'unscheduled'}"
        )
    else:
        print(f"Review: status={card['status']} not studied yet")
    print("\nSources:")
    for entry in conn.execute(
        """
        SELECT workbook_name, sheet_name, row_number, pos, meanings_json
        FROM source_entries
        WHERE word_id = ?
        ORDER BY band_rank, workbook_name, row_number
        """,
        (word["id"],),
    ):
        meanings = json.loads(entry["meanings_json"])
        print(
            f"- {entry['workbook_name']} / {entry['sheet_name']}{entry['row_number']}: "
            f"{entry['pos'] or 'pos ?'}"
        )
        for meaning in meanings[:5]:
            print(f"    {meaning}")
    return 0


def set_note(conn: sqlite3.Connection, term: str, note: str) -> int:
    word = lookup_word(conn, term)
    if word is None:
        print(f"No word found for: {term}")
        return 1
    conn.execute(
        """
        UPDATE study_cards
        SET notes = ?, updated_at = CURRENT_TIMESTAMP
        WHERE word_id = ?
        """,
        (note.strip(), word["id"]),
    )
    conn.commit()
    print(f"Saved note for {word['lemma']}")
    return 0


def choose_quiz_words(conn: sqlite3.Connection, limit: int) -> list[sqlite3.Row]:
    now = utc_now().isoformat()
    due = conn.execute(
        """
        SELECT words.id, words.lemma, words.best_band_label, words.best_band_rank, study_cards.status
        FROM words
        JOIN study_cards ON study_cards.word_id = words.id
        WHERE study_cards.next_review_at IS NOT NULL AND study_cards.next_review_at <= ?
        ORDER BY study_cards.next_review_at, words.best_band_rank, words.lemma
        LIMIT ?
        """,
        (now, limit),
    ).fetchall()
    if len(due) >= limit:
        return due
    need = limit - len(due)
    fresh = conn.execute(
        """
        SELECT words.id, words.lemma, words.best_band_label, words.best_band_rank, study_cards.status
        FROM words
        JOIN study_cards ON study_cards.word_id = words.id
        WHERE study_cards.last_reviewed_at IS NULL
        ORDER BY words.best_band_rank, words.lemma
        LIMIT ?
        """,
        (need,),
    ).fetchall()
    return list(due) + list(fresh)


def prompt_payload(conn: sqlite3.Connection, word_id: int, mode: str) -> tuple[str, str]:
    entries = conn.execute(
        """
        SELECT pos, meanings_json
        FROM source_entries
        WHERE word_id = ?
        ORDER BY band_rank, workbook_name, row_number
        """,
        (word_id,),
    ).fetchall()
    meanings = []
    parts = []
    for entry in entries:
        if entry["pos"] and entry["pos"] not in parts:
            parts.append(entry["pos"])
        for meaning in json.loads(entry["meanings_json"]):
            if meaning not in meanings:
                meanings.append(meaning)
    hint = " / ".join(parts) if parts else "unknown POS"
    if mode == "meaning-to-word":
        prompt = f"Meaning -> word\nPOS: {hint}\nClues:\n" + "\n".join(f"  - {m}" for m in meanings[:4])
        answer = conn.execute("SELECT lemma FROM words WHERE id = ?", (word_id,)).fetchone()["lemma"]
    else:
        lemma = conn.execute("SELECT lemma FROM words WHERE id = ?", (word_id,)).fetchone()["lemma"]
        prompt = f"Word -> meaning\n{lemma} ({hint})"
        answer = "\n".join(meanings[:4]) if meanings else "[No meanings recorded]"
    return prompt, answer


def apply_grade(conn: sqlite3.Connection, word_id: int, grade: str, prompt_mode: str) -> None:
    card = conn.execute("SELECT * FROM study_cards WHERE word_id = ?", (word_id,)).fetchone()
    now = utc_now()
    ease = float(card["ease"])
    interval = float(card["interval_days"])
    streak = int(card["streak"])
    correct_count = int(card["correct_count"])
    wrong_count = int(card["wrong_count"])

    if grade == "again":
        streak = 0
        wrong_count += 1
        interval = 1
        ease = max(1.3, ease - 0.2)
        status = "learning"
    elif grade == "hard":
        streak = max(1, streak)
        correct_count += 1
        interval = 2 if interval < 2 else interval * 1.2
        ease = max(1.3, ease - 0.15)
        status = "learning"
    elif grade == "good":
        streak += 1
        correct_count += 1
        interval = 3 if interval < 3 else interval * ease
        status = "review"
    else:
        streak += 1
        correct_count += 1
        interval = 5 if interval < 5 else interval * (ease + 0.3)
        ease = min(3.2, ease + 0.05)
        status = "review"

    next_review_at = (now + timedelta(days=interval)).isoformat()
    conn.execute(
        """
        UPDATE study_cards
        SET status = ?, correct_count = ?, wrong_count = ?, streak = ?, ease = ?,
            interval_days = ?, last_reviewed_at = ?, next_review_at = ?, updated_at = CURRENT_TIMESTAMP
        WHERE word_id = ?
        """,
        (
            status,
            correct_count,
            wrong_count,
            streak,
            ease,
            interval,
            now.isoformat(),
            next_review_at,
            word_id,
        ),
    )
    conn.execute(
        """
        INSERT INTO review_log (word_id, reviewed_at, prompt_mode, grade)
        VALUES (?, ?, ?, ?)
        """,
        (word_id, now.isoformat(), prompt_mode, grade),
    )
    conn.commit()


def run_quiz(conn: sqlite3.Connection, limit: int, mode: str) -> int:
    items = choose_quiz_words(conn, limit)
    if not items:
        print("No due or new cards found.")
        return 0
    random.shuffle(items)
    resolved_mode = mode
    print("Enter your answer, then self-grade with: again / hard / good / easy")
    for index, item in enumerate(items, start=1):
        if mode == "mixed":
            resolved_mode = random.choice(["meaning-to-word", "word-to-meaning"])
        prompt, answer = prompt_payload(conn, item["id"], resolved_mode)
        print(f"\n[{index}/{len(items)}] {prompt}")
        input("> ")
        print(f"Answer:\n{answer}")
        while True:
            grade = input("Grade [again/hard/good/easy]: ").strip().lower()
            if grade in {"again", "hard", "good", "easy"}:
                break
            print("Please type again, hard, good, or easy.")
        apply_grade(conn, item["id"], grade, resolved_mode)
    return 0


def list_due(conn: sqlite3.Connection, limit: int) -> None:
    now = utc_now().isoformat()
    rows = conn.execute(
        """
        SELECT words.lemma, words.best_band_label, study_cards.status, study_cards.next_review_at
        FROM words
        JOIN study_cards ON study_cards.word_id = words.id
        WHERE study_cards.next_review_at IS NOT NULL AND study_cards.next_review_at <= ?
        ORDER BY study_cards.next_review_at, words.best_band_rank, words.lemma
        LIMIT ?
        """,
        (now, limit),
    ).fetchall()
    if not rows:
        print("No cards are due right now.")
        return
    for row in rows:
        print(f"{row['lemma']:<20} {row['best_band_label']:<18} {row['status']:<8} {row['next_review_at']}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Personal Economist vocabulary database")
    parser.add_argument("--db", default=str(DEFAULT_DB_PATH), help="SQLite database path")
    subparsers = parser.add_subparsers(dest="command", required=True)

    import_parser = subparsers.add_parser("import", help="Import workbook data")
    import_parser.add_argument("--reset", action="store_true", help="Clear imported words and review history first")
    import_parser.add_argument("workbooks", nargs="*", help="Workbook paths (defaults to the five Economist files)")

    refresh_parser = subparsers.add_parser("refresh-source", help="Refresh workbook source data without rebuilding word ids")
    refresh_parser.add_argument("workbooks", nargs="*", help="Workbook paths (defaults to the five Economist files)")

    subparsers.add_parser("stats", help="Show database stats")

    search_parser = subparsers.add_parser("search", help="Show one word with all source entries")
    search_parser.add_argument("term")

    note_parser = subparsers.add_parser("note", help="Save a personal note for a word")
    note_parser.add_argument("term")
    note_parser.add_argument("note")

    due_parser = subparsers.add_parser("due", help="List cards due now")
    due_parser.add_argument("--limit", type=int, default=20)

    quiz_parser = subparsers.add_parser("quiz", help="Run an interactive quiz")
    quiz_parser.add_argument("--limit", type=int, default=10)
    quiz_parser.add_argument(
        "--mode",
        choices=["mixed", "meaning-to-word", "word-to-meaning"],
        default="mixed",
    )

    export_parser = subparsers.add_parser("export-enrichment-template", help="Export an Excel template for bulk enrichment")
    export_parser.add_argument("output", help="Output .xlsx file path")
    export_parser.add_argument("--band-rank", type=int, help="Only export one frequency band")
    export_parser.add_argument("--limit", type=int, help="Limit how many words to export")
    export_parser.add_argument("--missing-only", action="store_true", help="Only export words missing English definition or example sentence")

    import_enrichment_parser = subparsers.add_parser("import-enrichment", help="Import enrichment from .xlsx or .csv")
    import_enrichment_parser.add_argument("input", help="Input .xlsx or .csv file path")

    ai_parser = subparsers.add_parser("generate-enrichment-ai", help="Generate English definitions and examples with OpenAI")
    ai_parser.add_argument("--limit", type=int, default=20, help="How many words to generate in one run")
    ai_parser.add_argument("--band-rank", type=int, help="Only generate for one frequency band")

    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    db_path = Path(args.db)
    conn = connect(db_path)

    if args.command == "import":
        workbook_paths = [Path(p) for p in (args.workbooks or DEFAULT_WORKBOOKS)]
        missing = [str(path) for path in workbook_paths if not path.exists()]
        if missing:
            print("Missing workbook(s):")
            for item in missing:
                print(f"  {item}")
            return 1
        stats = import_workbooks(conn, workbook_paths, reset=args.reset)
        print(f"Imported into {db_path}")
        print(f"New words: {stats['inserted_words']}")
        print(f"New source entries: {stats['inserted_entries']}")
        print(f"Updated to better band: {stats['updated_best_band']}")
        return 0
    if args.command == "refresh-source":
        workbook_paths = [Path(p) for p in (args.workbooks or DEFAULT_WORKBOOKS)]
        missing = [str(path) for path in workbook_paths if not path.exists()]
        if missing:
            print("Missing workbook(s):")
            for item in missing:
                print(f"  {item}")
            return 1
        stats = refresh_workbooks(conn, workbook_paths)
        print(f"Refreshed source data in {db_path}")
        print(f"Deleted old source entries: {stats['deleted_entries']}")
        print(f"Inserted words: {stats['inserted_words']}")
        print(f"Inserted source entries: {stats['inserted_entries']}")
        print(f"Updated to better band: {stats['updated_best_band']}")
        print(f"Deleted orphan words: {stats['deleted_orphan_words']}")
        return 0
    if args.command == "stats":
        print_stats(conn)
        return 0
    if args.command == "search":
        return show_word(conn, args.term)
    if args.command == "note":
        return set_note(conn, args.term, args.note)
    if args.command == "due":
        list_due(conn, args.limit)
        return 0
    if args.command == "quiz":
        return run_quiz(conn, args.limit, args.mode)
    if args.command == "export-enrichment-template":
        from app.db import get_connection
        from app.enrichment_io import export_template

        web_conn = get_connection(db_path)
        count = export_template(
            web_conn,
            Path(args.output),
            band_rank=args.band_rank,
            limit=args.limit,
            missing_only=args.missing_only,
        )
        print(f"Exported {count} rows to {args.output}")
        return 0
    if args.command == "import-enrichment":
        from app.db import get_connection
        from app.enrichment_io import import_enrichment_rows, iter_import_rows

        input_path = Path(args.input)
        if not input_path.exists():
            print(f"File not found: {input_path}")
            return 1
        web_conn = get_connection(db_path)
        rows = iter_import_rows(input_path.name, input_path.read_bytes())
        stats = import_enrichment_rows(web_conn, rows)
        print(f"Imported enrichment from {input_path}")
        print(f"Updated: {stats['updated']}")
        print(f"Skipped blank rows: {stats['skipped']}")
        print(f"Missing words: {stats['missing_words']}")
        return 0
    if args.command == "generate-enrichment-ai":
        from app.db import get_connection
        from app.openai_enrichment import generate_enrichment_batch

        web_conn = get_connection(db_path)
        try:
            stats = generate_enrichment_batch(web_conn, limit=args.limit, band_rank=args.band_rank)
        except RuntimeError as exc:
            print(str(exc))
            return 1
        print(f"Selected: {stats['selected']}")
        print(f"Updated: {stats['updated']}")
        return 0
    parser.print_help()
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
